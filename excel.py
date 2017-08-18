from openpyxl import load_workbook
from collections import defaultdict

#filed = r"Z:\Packages\Breakdown Of Current Packages 08-04-2017.xlsx"
filed = r"Packages.xlsx"
wb = load_workbook(filed, data_only=True)
sh=wb["Current Packages"]
ws = wb.active

orl_pkgs = {}
lvn_pkgs = {}
nyc_pkgs = {}
haw_pkgs = {}
hhv_pkgs = {}
cal_pkgs = {}
sca_pkgs = {}
mtn_pkgs = {}
wdc_pkgs = {}

#KeyCodes - Outbound, Orlando, Las Vegas, Gold Mountain
def keycodes(m_row,mx_col,mx_row,x):
    num = 0
    for row in ws.iter_rows(min_row=m_row, max_col=mx_col, max_row=mx_row):
        for cell in row:
            num += 1
            if num == 10:
                num = 1
            if cell.value != None:
                print(cell.value)
                if num == 1:
                    pass
                    #listn.append(cell.value)
                #listn.append(str(num) + " val " + str(cell.value))
                #x[cell.value] = num
            #print(num)

            #print(cell.value)
            #x[cell.value] = None

price = []
day_res_discount = []
listn = []
package_terms = []

def get_pckg(x,y,pkg):
    for i in range(x,y):
        #print(sh['A%s' % i].value)
        listn.append(sh['A%s' % i].value)
        price.append(sh['E%s' % i].value)
        day_res_discount.append(sh['F%s' % i].value)
        package_terms.append(sh['N%s' % i].value)

get_pckg(5,80,orl_pkgs)
p = list(filter(None, listn))
q = list(filter(None, price))
r = list(filter(None, day_res_discount))
s = list(filter(None, day_res_discount))

p[:] = [item for item in p if item != "gifts - "]
p[:] = [item for item in p if item != "Gifts - "]


def ted():
    for f, b in zip(p,q):
        orl_pkgs[f] = str(b)

def sed():
    pass

new_dict = defaultdict(list)
def all_prce():
    for k, v in orl_pkgs.items():
        new_dict[v].append(k)

ted()
all_prce()

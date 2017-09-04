from openpyxl import load_workbook
from string import ascii_lowercase


wb = load_workbook('orgchart.xlsx', read_only=True)
sh = wb.worksheets[0]
ws = wb.active
column_header = {}
#Used Columns
Add = "Add/Delete/Change/Transfer/Rehire"
Add2 = "Add/Delete/Change"
Add3 = "Add/Delete/Change/Transfer"
windows = "Windows for Adds"
email = "Email Address"
Name = "AgentName"
cic_id = "CIC_ID"

#loop through and find column name
for c in ascii_lowercase:
	x = sh[c.upper() + "1"].value
	column_header[c.upper()] = sh[c.upper() + "1"].value

def orgchart_data(add, windows, agent_email,agent_name, agent_tsr):
    n = 2
    while n < sh.max_row :
        if sh[add + str(n)].value == "Add":
            username = sh[windows + str(n)].value
            agentName = sh[agent_name + str(n)].value
            email = sh[agent_email + str(n)].value
            tsr = sh[agent_tsr + str(n)].value
            
            print(username, email,agentName, tsr)
        n += 1

def sed(a):
     for k,v in column_header.items():
         if v == a:
             a = k
             return str(k)

def cheese():
    add = sed(Add) or sed(Add2) or sed(Add3)
    agent_username = sed(windows)
    agentemail = sed(email) or "AZ"
                                   
    agent_name =  sed(Name)
    agent_tsr = sed(cic_id)
  
    orgchart_data(add, agent_username, agentemail,agent_name, agent_tsr)
            
cheese()
 


from openpyxl import load_workbook
from string import ascii_lowercase
from tkinter import Tk, Label, Radiobutton,Button, W, NORMAL,DISABLED, StringVar
import sys
import os

wb = load_workbook('orgchart.xlsx', read_only=True)
sh = wb.worksheets[0]
ws = wb.active
column_header = {}
#Used Columns
Add = "Add/Delete/Change/Transfer/Rehire"
Add2 = "Add/Delete/Change"
Add3 = "Add/Delete/Change/Transfer"
windows = "Windows for Adds"
email = "Email Address"
Name = "AgentName"
cic_id = "CIC_ID"


#loop through and find column name
for c in ascii_lowercase:
	x = sh[c.upper() + "1"].value
	column_header[c.upper()] = sh[c.upper() + "1"].value

def orgchart_data(add, windows, agent_email,agent_name, agent_tsr):
    n = 2
    while n < sh.max_row :
        if sh[add + str(n)].value == "Add":
            username = sh[windows + str(n)].value
            agentName = sh[agent_name + str(n)].value
            email = sh[agent_email + str(n)].value
            tsr = sh[agent_tsr + str(n)].value
            
            print(username, email,agentName, tsr)
        n += 1

def sed(a):
     for k,v in column_header.items():
         if v == a:
             a = k
             return str(k)

def column_headers():
    add = sed(Add) or sed(Add2) or sed(Add3)
    agent_username = sed(windows)
    agentemail = sed(email) or "AZ"
                                   
    agent_name =  sed(Name)
    agent_tsr = sed(cic_id)
  
    orgchart_data(add, agent_username, agentemail,agent_name, agent_tsr)

class IAAutoGUI:
    def __init__(self, master, server = 0):
        self.master = master
        self.server = server
        
        master.title("IA Auto")
        self.master.minsize(width=300,height=200)
        
        self.serv_label = Label(master, text="Select Server")
        self.serv_label.grid(columnspan = 2, sticky = W)
        
        self.cms_button = Button(master, text="CMS", command=lambda:self.get_server(1))
        self.cms_button.grid(row = 1)

        self.sf_button = Button(master, text="Salesforce", command=lambda:self.get_server(2))
        self.sf_button.grid(row = 1, column = 1, sticky = W)

        self.loc_label = Label(master, text="Select Location")
        self.loc_label.grid(row = 2, columnspan = 2, sticky = W, pady = 10)

        self.orl_button = Button(master, text="orl", command=lambda:self.get_location("orl"),
                                 state=DISABLED)
        self.orl_button.grid(row = 3)
        
        self.lvn_button = Button(master, text="lvn", command=lambda:self.get_location("lvn"),
                                 state=DISABLED)
        self.lvn_button.grid(row = 3, column = 1, sticky = W )
        
        self.spg_button = Button(master,text="spg", command=lambda:self.get_location("spg"),
                                 state=DISABLED)
        self.spg_button.grid(row = 3, column = 2, sticky = W)      

        self.dept_label = Label(master, text="Choose Department")
        self.dept_label.grid(row = 4, sticky = W, pady = 10)

        self.department = StringVar()
        self.outbnd = Radiobutton(master,indicatoron = 0, width = 20, padx = 20,
                                  text='Outbound', variable=self.department, value = "outbnd")
        self.outbnd.grid(row = 5)

        self.ct = Radiobutton(master, indicatoron = 0, width = 20, padx = 20,
                              text='Call Transfer', variable=self.department, value = "ct")
        self.ct.grid(row = 5, column = 2)
        
        self.act = Radiobutton(master, indicatoron = 0, width = 20, padx = 20,
                               text='Activations', variable=self.department, value = "act")
        self.act.grid(row = 6)
        
        self.cc = Radiobutton(master, indicatoron = 0, width = 20, padx = 20,
                              text='Customer Care', variable=self.department, value = "cc")
        self.cc.grid(row = 6, column = 2)

        #self.outbnd.select()

        self.run_button =  Button(master, width = 10, height = 2, padx = 10,
                                  text = "Run", command = lambda:self.get_department())
        self.run_button.grid(row = 8, pady = 10)
        
        self.reset_button =  Button(master, width = 10, height = 2, padx = 10,
                                    text = "Reset", command = lambda:self.restart_program())
        self.reset_button.grid(row = 8, column =  2, pady = 10)

        col_count, row_count = root.grid_size()


        
    def get_server(self, button_id):
        if button_id == 1:
            self.server = 1
            print("CMS")
        elif button_id == 2:
            self.server = 2
            print("Salesforce")
            
        self.cms_button.configure(state=DISABLED)
        self.sf_button.configure(state=DISABLED)
        self.orl_button.configure(state=NORMAL)
        self.lvn_button.configure(state=NORMAL)
        self.spg_button.configure(state=NORMAL)
    
    def get_location(self, location):
        self.location = location
        
        print(location)
        self.orl_button.configure(state=DISABLED)
        self.lvn_button.configure(state=DISABLED)
        self.spg_button.configure(state=DISABLED)

        if self.server == 1:
            if location == "spg":
                self.outbnd.configure(state = DISABLED)
                self.act.configure(state = DISABLED)
                self.cc.configure(state = DISABLED)
            elif location == "lvn":
                self.act.configure(state = DISABLED)
                self.cc.configure(state = DISABLED)
            else:
                pass
        else:
            self.ct.configure(state = DISABLED)
            self.act.configure(state = DISABLED)
            self.cc.configure(state = DISABLED)
            self.act.configure(state = DISABLED)
            self.cc.configure(state = DISABLED)
            
    def get_department(self):
        self.department = self.department.get()
        print(self.department)
        
        self.outbnd.configure(state = DISABLED)
        self.ct.configure(state = DISABLED)
        self.act.configure(state = DISABLED)
        self.cc.configure(state = DISABLED)
        self.act.configure(state = DISABLED)
        self.cc.configure(state = DISABLED)
        column_headers()
        
    def restart_program(self):
        python = sys.executable
        os.execl(python, python, * sys.argv)
        
            
        

if __name__ == '__main__':
    root = Tk()
    my_gui = IAAutoGUI(root)
    root.mainloop()

print(my_gui.server)
print(my_gui.location)
print(my_gui.department)

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from string import ascii_lowercase
import time

user_name = ""
passwd = ""

driver = webdriver.Firefox()
by_name = driver.find_element_by_name
by_id = driver.find_element_by_id

wb= load_workbook('OrlandoOutbound.xlsx', read_only=True)
sh = wb.worksheets[1]
ws = wb.active
column_header = {}

def get_alphabet():
    """Loop Through Alphabet and Column Headers in Excel"""
    for c in ascii_lowercase:
        _x = sh[c.upper() + "1"].value
        column_header[c.upper()] = sh[c.upper() + "1"].value

def get_header(getLetter):
    """Match Header of File"""
    for k, v in column_header.items():
        if v == getLetter:
            getLetter = k
            return str(k)

def column_headers():
    """Assign ColumnHeader Names"""
    Add = "Add/Delete/Change/Transfer/Rehire"
    Add2 = "Add/Delete/Change"
    Add3 = "Add/Deactivate/Reactivate/Change"
    Add4 = "Add/Delete/Change/Transfer"
    windows = "Windows Username"
    windows2 = "Windows for Adds"
    Name = "AgentName"
    cic_id = "CIC_ID"
    cic_id2 = "CIC ID"
    queue = "Campaign"
    Shift = "Shift"

    add = get_header(Add) or get_header(Add2) or get_header(Add3) or get_header(Add4)
    agent_username = get_header(windows) or get_header(windows2) 
    agent_name = get_header(Name)
    agent_tsr = get_header(cic_id) or get_header(cic_id2)
    campaign = get_header(queue)
    shift = get_header(Shift)
    
    orgchart_data(add, agent_username, agent_name, agent_tsr, campaign, shift)

def BrowserSetup():
    driver.get("https://identitynow.hgv.com/")
    #driver.get("https://hgv.my.salesforce.com/home/home.jsp?tsid=02u38000000NcOF")
    #assert "HGV IdentityNow" in driver.title
    username = by_id('username')
    password = by_id('password')
    username.send_keys(user_name)
    password.send_keys(passwd)

def startUp():
    by_id("signIn").click()
    time.sleep(5)
    driver.get("https://hgv.my.salesforce.com/home/home.jsp?tsid=02u38000000NcOF")
    time.sleep(7)
    by_id("01r38000000Hv1x_Tab").click()
    time.sleep(2)
    select = Select(driver.find_element_by_id("fcf"))
    select.select_by_visible_text("Available & Callable - Cell - Pacific")
    by_name("go").click()
    time.sleep(4)

def leads():
    by_id("00B38000007Xnvh_paginator_rpp_target").click()
    driver.find_element_by_xpath("//*[text()='200']").click()
    time.sleep(1)
    by_id("allBox").click()
    driver.find_element_by_xpath("//input[@type='submit' and @value='Change Owner']").click()

def orgchart_data(add, windows, agent_name, agent_tsr, campaign, shift):
    n = 2
    while n < sh.max_row + 1:
        username = sh[windows + str(n)].value
        agentName = sh[agent_name + str(n)].value
        tsr = sh[agent_tsr + str(n)].value
        user_campaign = sh[campaign + str(n)].value
        agent_shift = sh[shift + str(n)].value
        if sh[agent_name + str(n)].value != None and sh[add + str(n)].value != "Deactivate":
            #print(agentName, tsr, user_campaign, agent_shift)
            if sh[campaign + str(n)].value != "New Hire-SF Cell Phone2":
                AssignUser(agentName)
                time.sleep(3)
                SelectQueue()
                time.sleep(2)
                SelectLeads()          
        n += 1

def GetUser():
    pass

def SelectQueue():
    select.select_by_visible_text("Available & Callable - Cell - Pacific")
    by_name("go").click()
    time.sleep(4)

def AssignUser(AgentName):
    time.sleep(3)
    user = by_id('newOwn')
    user.send_keys(AgentName) #Get User Name
    time.sleep(0.3)
    by_name('cancel').click()

def SelectLeads():
    by_id("00B38000007Xnvh_paginator_rpp_target").click()
    driver.find_element_by_xpath("//*[text()='200']").click()
    time.sleep(1)
    by_id("allBox").click()
    driver.find_element_by_xpath("//input[@type='submit' and @value='Change Owner']").click()

def main():
    BrowserSetup()
    startUp()
    leads()
    get_alphabet()
    column_headers()

main()
#driver.close()

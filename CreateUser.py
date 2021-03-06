"""
Automates Interaction Administrator
"""
#import time
from string import ascii_lowercase
import pywinauto
#from pywinauto import application
#from pywinauto.keyboard import SendKeys
#from pywinauto import Desktop
from pywinauto.application import Application
#from pywinauto.findwindows import find_window
from openpyxl import load_workbook

#CMS
flname = 'excel_orgchart\HGV_OrgChart_Template_Orlando_Outbound_Salesforce_03262018.xlsx'
shName = input("Enter Sheet: ")
sheetName = int(shName) - 1

orl_outbnd_cms = ["MKT-Outbound-Callback", "MKT-Outbound-Main2", "Orl_OUT_SUP"]
lvn_outbnd_cms = ["LAS_OUT_SUP", "MKT-Outbound-Callback", "MKT-Outbound-Main2"]
spg_ct_cms = ["LOC-SPG-MKT-HRCC", "MKT-InbCT-Callback", "MKT-InbCT-HRCC"]

orl_ct_cms = [
    "CT Priority 1", "CT Priority 2", "LOC-ORL-MKT-HRCC",
    "MKT-InbCT-Callback", "MKT-InbCT-HRCC",
    ]
lvn_ct_cms = [
    "CT Priority 1", "CT Priority 2", "LOC-LAS-MKT-HRCC",
    "MKT-InbCT-Callback", "MKT-InbCT-HRCC",
    ]
orl_act_cms = [
    "LOC-ORL-MKT-ACT", "MKT-Activations-CallBack", "MKT-ACT-Main",
    ]
orl_cc_cms = [
    "LOC-ORL-MKT-CC", "MKT-CC-BookDates", "MKT-CC-BookDates-Priority1",
    "MKT-CC-CustomerService", "MKT-CC-CustomerService-Priority2",
    ]
orl_own_cms = [
    "LOC-ORL-MKT-OWN", "MKT-Owner-Extension1", "MKT-Owner-Extension2", "MKT-Owners-Callback",
    "MKT-Owners-LeadGen1", "MKT-Owners-LeadGen2", "MKT-Owners-Main", "MKT-Owners-Main2",
    "MKT-Owners-NonResponder"
    ]

wrkqueues = {
    "orl_outbnd_cms":orl_outbnd_cms, "orl_ct_cms":orl_ct_cms, "orl_act_cms":orl_act_cms,
    "orl_cc_cms":orl_cc_cms, "spg_ct_cms":spg_ct_cms, "lvn_outbnd_cms":lvn_outbnd_cms,
    "lvn_ct_cms":lvn_ct_cms,
    }

#Manual
orl_outbnd_sf = ["LOC-ORL-MKT-SalesForce", "SF-Orlando-Manual", "SF-RestrictDialing"]
spg_outbnd_sf = ["LOC-SPG-MKT-SalesForce", "SF-Springfield-Manual", "SF-RestrictDialing"]
lvn_outbnd_sf = ["LOC-LAS-MKT-SalesForce", "SF-Vegas-Manual", "SF-RestrictDialing"]

licenses = [
    "Interaction Optimizer Client Access", "Interaction Optimizer Real-time Adherence Tracking",
    "Interaction Optimizer Schedulable",
    ]

roles = ["MKT-Agent", "MKT-SF-Agent", "MKT-CC-Agent"]

column_header = {}

serv = input("1:CMS\n2:Salesforce\nSelect Option:")
serv = int(serv)

location = input("Location orl, spg, lvn:")
location = location.lower()

if serv == 1:
    if location == "spg":
        department = input("Select a department - ct:  ")
    elif location == "lvn":
        department = input("Select a department - outbnd, ct: ")
    else:
        department = input("Select a department - outbnd, ct, act, cc: ")
    department = department.lower()

wb = load_workbook(flname, read_only=True)
sh = wb.worksheets[sheetName]
ws = wb.active

app = Application(backend='uia')

if serv == 1:
    p = pywinauto.findwindows.find_element(title="Interaction Administrator - [HiltonACD]")
else:
    p = pywinauto.findwindows.find_element(title="Interaction Administrator - [HiltonTCPA]")
app.connect(handle=p.handle)
if serv == 1:
    dlg = app.window(title="Interaction Administrator - [HiltonACD]")
else:
    dlg = app.window(title="Interaction Administrator - [HiltonTCPA]")
typein = app.dlg.type_keys

#app.dlg.print_control_identifiers() #Check Identifiers

def main():
    """Main Function"""
    get_alphabet()
    column_headers()

def get_alphabet():
    """Loop Through Alphabet and Column Headers in Excel"""
    for c in ascii_lowercase:
        _x = sh[c.upper() + "1"].value
        column_header[c.upper()] = sh[c.upper() + "1"].value

def get_header(getLetter):
    """Match Header of File"""
    for k, v in column_header.items():
        if v == getLetter:
            getLetter = k
            return str(k)

def column_headers():
    """Assign ColumnHeader Names"""
    Add = "Add/Delete/Change/Transfer/Rehire"
    Add2 = "Add/Delete/Change"
    Add3 = "Add/Deactivate/Reactivate/Change"
    Add4 = "Add/Delete/Change/Transfer"
    windows = "Windows Username"
    windows2 = "Windows for Adds"
    Name = "AgentName"
    cic_id = "CIC_ID"
    cic_id2 = "CIC ID"

    add = get_header(Add) or get_header(Add2) or get_header(Add3) or get_header(Add4)
    agent_username = get_header(windows) or get_header(windows2) 
    agent_name = get_header(Name)
    agent_tsr = get_header(cic_id) or get_header(cic_id2)

    orgchart_data(add, agent_username, agent_name, agent_tsr)

def orgchart_data(add, windows, agent_name, agent_tsr):
    n = 2
    while n < sh.max_row + 1:
        if sh[add + str(n)].value == "Add":
            username = sh[windows + str(n)].value
            agentName = sh[agent_name + str(n)].value
            tsr = sh[agent_tsr + str(n)].value

            print("Adding User - " + username, agentName, tsr)

            try:
                config(tsr, username)
                get_user_details(agentName)
                auto_acd()
                if serv == 1:
                    cms_roles(department)
                    get_cms_workgroups(location, department)
                    if department == "ct" or department == "cc":
                        licensing()
                    app.dlg.Ok.click_input() #Change When Done
                else:
                    sf_roles()
                    get_sf_workgroups()
                    app.dlg.Ok.click_input() #Change When Done
            except:
                if app.dlg["A User with that name already exists"].exists() == True:
                    print("User Already Exists " + username, agentName, tsr)
                    app.dlg.OK.click_input()
                    app.dlg.Cancel.click_input()
        n += 1

def config(tsr, win_username, passwd="10102015", domain="hgvcnt\\"):
    """Assign Configuration"""
    app.dlg.type_keys('^n')
    app.dlg.Edit0.type_keys(str(tsr) + "{ENTER}")
    app.dlg.Edit3.type_keys(passwd) #Password
    app.dlg.Edit4.type_keys(passwd) #Confirm Password
    app.dlg.Edit7.type_keys(domain + win_username) #Domain User
    get_location()

def get_location():
    """Gets Location from combobox"""
    app.dlg["ComboBox4"].click_input()
    if location == "orl":
        app.dlg["Orlando - Metro Center"].select()
    if location == "spg":
        app.dlg["Spingfield"].select()
    if location == "lvn":
        app.dlg["Las Vegas"].select()

def get_user_details(agentName):
    """Assign Personal Info"""
    name = agentName.split(" ")

    app.dlg.TabItem3.click_input()
    app.dlg.Edit0.type_keys(name[0]) #FirstName
    app.dlg.Edit2.type_keys(name[1]) #LastName
    app.dlg.Edit4.type_keys(name[0] + "{SPACE}" + name[1]) #Display Name

def auto_acd():
    """Assign Auto-ACD"""
    app.dlg.ACD.click_input()
    app.dlg.ListItem3.click_input()
    app.dlg.CheckBox0.click_input()

def cms_roles(deptmnt):
    """Assign Roles for cms"""
    app.dlg.Roles.click_input()
    app.dlg.Add.click_input()

    if deptmnt == "ct" or deptmnt == "outbnd":
        app.dlg["MKT-Agent"].select()
    elif deptmnt == "act" or deptmnt == "cc":
        app.dlg["MKT-CC-Agent"].select()

    app.dlg.OK.click_input()

def sf_roles():
    """Assign Roles in SalesForce"""
    app.dlg.Roles.click_input()
    app.dlg.Add.click_input()
    app.dlg.ListItem6.select()
    app.dlg.OK.click_input()

def get_cms_workgroups(loc, dept):
    """Assign CMS WorkGroups based on department"""
    if location == loc:
        if department == dept:
            agent_cms_workgroups(wrkqueues[loc + "_" + dept + "_" + "cms"])
            
def get_sf_workgroups(): #helper function here
    """Assign Salesforce WorkGroups based on Selection"""
    if location == "orl":
        agent_sf_workgroups(orl_outbnd_sf)
    if location == "spg":
        agent_sf_workgroups(spg_outbnd_sf)
    if location == "lvn":
        agent_sf_workgroups(lvn_outbnd_sf)

def agent_cms_workgroups(wrkgrps):
    """Assign CMS WorkGroups"""
    num = 0
    app.dlg.Workgroups.click_input()
    app.dlg.OK.click_input()

    for x in wrkgrps:
        if location == "orl":
            if department == "ct":
                LstBoxAdd(x)

            if department == "act":
                LstBoxAdd(x)
            if department == "cc":
                LstBoxAdd(x)

        if location == "spg":
            LstBoxAdd(x)
            
        if location == "lvn":
            if department == "outbnd":
                LstBoxAdd(x)
            if department == "ct":
                LstBoxAdd(x)
    
def agent_sf_workgroups(wrkgrps):
    """Assign SF WorkGroups"""
    num = 0
    app.dlg.Workgroups.click_input()
    app.dlg.OK.click_input()

    for x in wrkgrps:
        if location == "orl":
            LstBoxAdd(x)
        if location == "spg":
            LstBoxAdd(x)
        if location == "lvn":
            LstBoxAdd(x)

def LstBoxAdd(x):
    app.dlg['ListBox'].click_input()
    app.dlg['Listbox'].type_keys(x)
    app.dlg.Add.click_input()

def licensing():
    """Assign Licensing"""
    app.dlg.Licensing.click_input()
    app.dlg.OK.click_input()
    app.dlg['Enable Licenses'].click_input()
    for ls in licenses:
        app.dlg[ls].type_keys("{SPACE}")

main()

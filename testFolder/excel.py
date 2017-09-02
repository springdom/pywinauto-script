from openpyxl import load_workbook
from collections import defaultdict
import json

#filed = r"Z:\Packages\Breakdown Of Current Packages 08-04-2017.xlsx"
filed = r"Packages.xlsx"
wb = load_workbook(filed, data_only=True)
sh=wb["Current Packages"]
ws = wb.active

orl_pkgs = {}

"""
#KeyCodes - Outbound, Orlando, Las Vegas, Gold Mountain
def keycodes(m_row,mx_col,mx_row,x):
    num = 0
    for row in ws.iter_rows(min_row=m_row, max_col=mx_col, max_row=mx_row):
        for cell in row:
            num += 1
            if num == 10:
                num = 1
            if cell.value != None:
                print(cell.value)
                if num == 1:
                    pass
                    #listn.append(cell.value)
                #listn.append(str(num) + " val " + str(cell.value))
                #x[cell.value] = num
            #print(num)

            #print(cell.value)
            #x[cell.value] = None
"""

#price = []
pckgs = []

for i in range(11, 1):
    ws.cell(row=i, column=2).value = i

def get_pckg(x,y,pkg):
    for i in range(x,y):
        pckg_code = str(sh['A%s' % i].value)
        price = str(sh['E%s' % i].value)
        day_res_discount = sh['F%s' % i].value
        if pckg_code != "None" or price != "None":
            if pckg_code != "gifts - " and pckg_code != "Gifts - ":
                pckgs.append({'Package Code' : pckg_code,
                              'Price' : price,
                        })
        #print(sh['A%s' % i].value)
       

get_pckg(5,80,orl_pkgs)
print(pckgs[0]['Package Code'])
"""
print(json.dumps(pckgs, sort_keys=True,
                 indent=4,separators=(',',': ')))
"""

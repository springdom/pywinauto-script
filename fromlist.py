import time
import win32clipboard
import pywinauto
from pywinauto import application
from pywinauto.keyboard import SendKeys
from pywinauto import Desktop
from pywinauto.application import Application
from pywinauto.findwindows import find_window
from openpyxl import load_workbook
from openpyxl import load_workbook

wb = load_workbook('changetokick.xlsx', data_only=True)
sh=wb["Sheet1"]
ws = wb.active

def keycodes(m_row,mx_col,mx_row):
    for row in ws.iter_rows(min_row=m_row, max_col=mx_col, max_row=mx_row):
        for cell in row:
                #print(cell.value)
                lead = cell.value.split("-")
                #print(leadvar)
                change_to_kick(lead[1],lead[0])

def set_window():
    app = Application(backend='uia')
    p = pywinauto.findwindows.find_element(best_match="Notepad")
    app.connect(handle=p.handle)
    dlg = app.window(best_match="Notepad")
    app.dlg.set_focus()
    #app.dlg.type_keys("Test")


def change_to_kick(leadn, loca):

    set_window()
    SendKeys(leadn + '{ENTER}')
    SendKeys(loca + '{ENTER}')
    SendKeys('{ENTER 2}')
    
    package_sel = input("Enter Package Number: ")
    
    set_window()
    SendKeys(package_sel + '{ENTER}')
    SendKeys("2")
    SendKeys("{ENTER}")
    SendKeys("9")
    SendKeys("{ENTER 2}")
    SendKeys("f" + "{ENTER}")
    

keycodes(3,1,11)

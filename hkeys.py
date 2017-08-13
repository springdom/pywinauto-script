import time
import pywinauto
from pywinauto.keyboard import SendKeys
from pywinauto.application import Application
from openpyxl import load_workbook
#from pywinauto import application
#import win32clipboard
#from pywinauto import Desktop
#from pywinauto.findwindows import find_window

#app.dlg.print_control_identifiers() #Check Identifiers

app = Application(backend='uia')
p = pywinauto.findwindows.find_element(best_match="AbsoluteTelnet")
#p = pywinauto.findwindows.find_element(best_match="Notepad")
app.connect(handle=p.handle)
dlg = app.window(best_match="AbsoluteTelnet")
#dlg = app.window(best_match="Notepad")

wb_keycodes = load_workbook('KeyCodes.xlsx', data_only=True)
sh_keycodes = wb_keycodes["Sheet1"]
ws_keycodes = wb_keycodes.active

wb_leadkick = load_workbook('changetokick.xlsx', data_only=True)
sh_leadkick = wb_leadkick["Sheet1"]
ws_leadkick = wb_leadkick.active

outbnd_orl = {}
orl_call_trsf = {}
lvn_call_trsf = {}
gldmtn_call_trsf = {}

opt1 = [1, 2, 1, 2]
opt2 = [1, 2, 1, 15, 4, 8]
opt3 = [1, 2, 1, 15, 4, 10]

typein = app.dlg.type_keys

def main():
    print("-" * 20)
    print("Select From Menu")
    print("0: Check Availability")
    print("8: Check Tours")
    print("7: Cashback Check")
    print("-" * 20)
    print("1: Book Dates")
    print("2: Unlock Lead")
    print("-" * 20)
    print("3: Kick Package")
    print("4: Change to Kick")
    print("5: Create Package")
    print("-" * 20)
    Sel = int(input("Select Option: "))
    print("")

    if Sel >= 1 and Sel <= 3 or Sel == 7:
        try:
            lead = input("Enter Lead: ")
            loc = input("Enter Location: ")
        except (KeyboardInterrupt, SystemExit):
            main()

    if Sel == 1:
        lead_package(lead, loc)
    elif Sel == 2:
        unlock_lead(lead, loc)
    elif Sel == 3:
        kick_package(lead, loc)
    elif Sel == 4:
        """
        win32clipboard.OpenClipboard()
        data = win32clipboard.GetClipboardData()
        win32clipboard.CloseClipboard()
        leadvar = data.split("-")
        change_to_kick(leadvar[1],leadvar[0])
        """
        kick_list(3, 1, 56)
    elif Sel == 5:
        build_package()
    elif Sel == 0:
        check_dates()
    elif Sel == 8:
        check_tours()
    elif Sel == 7:
        check_cashback(lead, loc)
    else:
        print("Invalid Option")
        main()


#Window Settings
"""
def set_window():
    app.dlg.set_focus()
    #app.dlg.type_keys("Test" + "{ENTER}")
"""
#Enter Mode
def mode(opt):
    typein('{BS 10}')
    typein('{ENTER}')
    time.sleep(1.5)

    for num in opt:
        typein(str(num))
        typein('{ENTER}')
    typein('{ENTER}')

#KeyCodes - Outbound, Orlando, Las Vegas, Gold Mountain
def keycodes(m_row, mx_col, mx_row, x):
    num = 0
    for row in ws_keycodes.iter_rows(min_row=m_row, max_col=mx_col, max_row=mx_row):
        for cell in row:
            num += 1
            if num == 1:
                #print(cell.value)
                x[cell.value.lower()] = None
            if m_row > 17:
                if num == 33:
                    num = 0
            else:
                if num == 3:
                    num = 0

def assign_key(n, x):
    if n > 17:
        for k, v in x.items():
            x[k] = sh_keycodes['D%s' % n].value
            n += 1
    else:
        for k, v in x.items():
            x[k] = sh_keycodes['B%s' % n].value
            n += 1

def kick_list(m_row, mx_col, mx_row):
    for row in ws_leadkick.iter_rows(min_row=m_row, max_col=mx_col, max_row=mx_row):
        for cell in row:
            lead = cell.value.split("-")
            change_to_kick(lead[1], lead[0])

#Lead Mktg Package Entry & Edit
def lead_package(leadn, loca):
    region = input("Enter Region: ")
    arrival = input("Enter Arrival: ")
    nights = input("Enter Nights: ")

    mode(opt1)

    typein('{TAB}' + '2' + '{ENTER}')
    typein(region + '{ENTER 2}')
    typein(arrival + '{ENTER 2}')
    typein(nights)
    typein('{TAB}' + '3' + '{ENTER}')
    
    checkavail = input("Available: y or n: ")
    if checkavail == "n":
        main()

    mode(opt1)

    typein(leadn + '{ENTER}')
    typein(loca + '{ENTER}')
    typein('{ENTER 2}')

    try:
        package_sel = input("Select Package Number: ")

        #set_window()

        typein('{ENTER}')
        typein(package_sel + '{ENTER}')
        typein('7' + '{ENTER}' + 'A' + '{ENTER}')
        propn = input("Enter Property Number: ")
        unitt = input("Enter Unit Type: ")
        adults = input("Adults: ")
        kids = input("Kids: ")
        #arrival = input("Arrival: ")
        #nights = input("Nights: ")

        #set_window()

        typein(propn + '{ENTER 2}')
        typein(unitt + '{ENTER 4}')
        typein(adults + '{ENTER 2}')
        typein(kids + '{ENTER 2}')
        typein(arrival + '{ENTER}')
        typein(nights + '{ENTER 2}')
        bckspace = input("Go back? y or n: ") # Backspace

        #set_window()

        if bckspace == "y":
            typein('{BS}')

        proceed = input("Proceed - y or n: ")

        #set_window()

        if proceed == 'y':
            typein('{ENTER}')

        time.sleep(1)
        office = input("Enter Office: ")
        manifest = input("Manifest: ")
        #set_window()
        typein('{TAB}' + office + '{ENTER}' + '{ENTER}' + '{ENTER 2}')
        toursa = input("Tours Available y or n: ")

        #set_window()

        if toursa == "y":
            typein('1' + '{ENTER 2}')
            typein(manifest)
            typein('{ENTER 9}')
            time.sleep(1.30)
            typein('{ENTER 17}')
            typein('f' + '{ENTER}')
            typein('y' + "{ENTER}")
            amount = input("Enter amount: ")
            #set_window()
            typein(amount)
            typein("{ENTER}" + "n")
        else:
            pass
        main()
    except (KeyboardInterrupt, SystemExit):
        main()

#Unlock Lead
def unlock_lead(leadn, loca):
    mode(opt2)
    #set_window()
    typein(leadn + '{ENTER}')
    typein(loca + '{ENTER}')
    typein('{ENTER 6}')
    main()

#Kick Package
def kick_package(leadn, loca):
    mode(opt1)
    txtcmnt = "Package Kicked as requested"
    typein(leadn + '{ENTER}')
    typein(loca + '{ENTER}')
    typein('{ENTER 2}')
    package_sel = input("Select Package Number: ")

    #set_window()

    typein('{ENTER}')
    typein(package_sel + '{ENTER}')
    typein('20' + '{ENTER}')
    typein('2' + '{ENTER}')
    typein('14' + '{ENTER}')
    typein('{TAB}')
    time.sleep(3)
    typein('{ENTER 10}')
    typein('13')
    #typein('A')
    #typein('10' + '{ENTER 5}')
    #typein(txtcmnt.replace(" ","{SPACE}") + "{ENTER}")
    main()

#Change To Kick
def change_to_kick(leadn, loca):
    mode(opt3)

    typein(leadn + '{ENTER}')
    typein(loca + '{ENTER}')
    typein('{ENTER 2}')

    package_sel = input("Enter Package Number: ")
    #set_window()

    typein(package_sel + '{ENTER}')
    typein("2")
    typein("{ENTER}")
    typein("9")
    typein("{ENTER 2}")
    typein("f" + "{ENTER}")
    #main()

#Check Dates
def check_dates():
    region = input("Enter Region: ")
    arrival = input("Enter Arrival: ")

    mode(opt1)

    typein('{TAB}' + '2' + '{ENTER}')
    typein(region + '{ENTER 2}')
    typein(arrival + '{ENTER}')
    typein('{TAB}' + '2' + '{ENTER}')
    main()

#Check Tours
def check_tours():
    office = input("Enter Office: ")
    starting = input("Enter Tour Date: ")

    mode(opt1)

    typein('{TAB}' + '3' + '{ENTER}')
    typein(office + '{ENTER}')
    typein(starting + '{ENTER 3}')
    typein('{TAB}' + '2' + '{ENTER}')
    main()

def check_cashback(leadn, loca):
    mode(opt3)

    typein(leadn + '{ENTER}')
    typein(loca + '{ENTER}')
    typein('{ENTER 2}')
    package_sel = input("Enter Package Number: ")
    #set_window()
    typein(package_sel + '{ENTER}')
    typein('18' + '{ENTER}')
    time.sleep(1)
    typein('{BS 3}')
    main()

#Build A Package
def build_package():
    mode(opt1)
    #Check if Exist first
    typein("{ENTER}")
    last_name = input("Enter Last Name: ")
    first_name = input("Enter First Name: ")
    phone_no = input("Phone Number: ")
    phone_no2 = input("Second Phone Number: ")
    country = input("Enter Country Code: ")
    addr = input("Address: ")
    zip_code = input("Input Zip Code: ")
    hhn = input("Hilton Honors Number: ")
    email = input("Enter Email: ")
    lang = input("Preferred Language: ")
    natio = input("Nationality: ")

    #set_window()

    #last_name,first_name,Phone Number
    typein(last_name.replace(" ", "{SPACE}") + "{ENTER}")
    typein(first_name.replace(" ", "{SPACE}") + "{ENTER}")
    typein(phone_no + "{ENTER 2}")
    #country,Enter*4
    if "/" in last_name or "/" in first_name:
        typein(country + "{ENTER 5}")
    else:
        typein(country + "{ENTER 4}")
    typein(zip_code + "{ENTER}")
    typein(addr.replace(" ", "{SPACE}") + "{ENTER 5}")
    time.sleep(2)

    #Check if address good
    addrgood = input("Select Option: override continue 'o or c': ")

    #set_window()
    if addrgood == "o":
        typein("o" + "{ENTER}")
    elif addr == "c":
        pass

    typein(phone_no2)
    typein("{ENTER 9}")
    typein(hhn + "{ENTER 2}")
    typein(email + "{ENTER 2}")
    typein(lang + "{ENTER}")
    typein(natio + "{ENTER 2}")

    ###Marketing Key Section
    ##Check Department
    dptm = input("Department outbnd, orl, gldmtn, lvn: ")
    key_loca = input("Location: ")
    promo = input("POS promo: ")
    pckgcode = input("Package Code: ")
    stt = input("Marital Status: ")
    tsr = input("Agent tsr: ")
    if dptm == "outbnd":
        keycodes(4, 3, 12, outbnd_orl)
        assign_key(4, outbnd_orl)
        mktkey = outbnd_orl[key_loca]
    if dptm == "orl":
        keycodes(18, 33, 27, orl_call_trsf)
        assign_key(18, orl_call_trsf)
        mktkey = orl_call_trsf[key_loca]
    if dptm == "gldmtn":
        keycodes(46, 33, 55, gldmtn_call_trsf)
        assign_key(46, gldmtn_call_trsf)
        mktkey = gldmtn_call_trsf[key_loca]
    if dptm == "lvn":
        keycodes(32, 33, 41, lvn_call_trsf)
        assign_key(32, lvn_call_trsf)
        mktkey = lvn_call_trsf[key_loca]

    #set_window()
    typein(mktkey + "{ENTER}")
    typein("{BS}")
    typein("{ENTER 2}")
    typein(promo + "{ENTER 6}")

    #Payment
    typein(pckgcode + "{ENTER 5}")
    price = input("Price: ")
    #set_window()
    typein(str(price) + "{ENTER 3}")
    ##add CC
    #is cc good then
    ccgood = input("CC accepted: ")

    if not ccgood:
        main()

    #set_window()
    typein("{ENTER 2}")
    typein(str(tsr) + "{ENTER}")
    typein(str(tsr) + "{ENTER}")
    typein(str(tsr) + "{ENTER 12}")

    typein("1" + "{ENTER}")
    typein("1" + "{ENTER}")
    typein(mktkey + "{ENTER 3}")
    typein(stt + "{ENTER 10}")
    main()

main()

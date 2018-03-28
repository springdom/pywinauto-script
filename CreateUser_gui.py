"""
Automates Interaction Administrator
"""
#import time
from string import ascii_lowercase
import pywinauto
from pywinauto.application import Application
from openpyxl import load_workbook
from tkinter import Tk, Label, Radiobutton,Button, W, NORMAL,DISABLED, StringVar, ttk
from tkinter import filedialog as fd
import sys

#CMS
orl_outbnd_cms = ["MKT-Outbound-Callback", "MKT-Outbound-Main2", "Orl_OUT_SUP"]
lvn_outbnd_cms = ["LAS_OUT_SUP", "MKT-Outbound-Callback", "MKT-Outbound-Main2"]
spg_ct_cms = ["LOC-SPG-MKT-HRCC", "MKT-InbCT-Callback", "MKT-InbCT-HRCC"]

orl_ct_cms = [
    "CT Priority 1", "CT Priority 2", "LOC-ORL-MKT-HRCC",
    "MKT-InbCT-Callback", "MKT-InbCT-HRCC",
    ]
lvn_ct_cms = [
    "CT Priority 1", "CT Priority 2", "LOC-LAS-MKT-HRCC",
    "MKT-InbCT-Callback", "MKT-InbCT-HRCC",
    ]
orl_act_cms = [
    "LOC-ORL-MKT-ACT", "MKT-Activations-CallBack", "MKT-ACT-Main",
    "MKT-CC-BookDates", "MKT-CC-BookDates-Priority1", "MKT-CC-CustomerService",
    "MKT-CC-CustomerService-Priority2",
    ]
orl_cc_cms = [
    "LOC-ORL-MKT-CC", "MKT-CC-BookDates", "MKT-CC-BookDates-Priority1",
    "MKT-CC-CustomerService", "MKT-CC-CustomerService-Priority2",
    ]

wrkqueues = {
    "orl_outbnd_cms":orl_outbnd_cms, "orl_ct_cms":orl_ct_cms, "orl_act_cms":orl_act_cms,
    "orl_cc_cms":orl_cc_cms, "spg_ct_cms":spg_ct_cms, "lvn_outbnd_cms":lvn_outbnd_cms,
    "lvn_ct_cms":lvn_ct_cms,
    }

#SalesForce
orl_outbnd_sf = ["LOC-ORL-MKT-SalesForce", "SF-Orlando-Manual", "SF-RestrictDialing"] #Manual
spg_outbnd_sf = ["LOC-SPG-MKT-SalesForce", "SF-Springfield-Manual", "SF-RestrictDialing"]
lvn_outbnd_sf = ["LOC-LAS-MKT-SalesForce", "SF-Vegas-Manual", "SF-RestrictDialing"]

licenses = [
    "Interaction Optimizer Client Access", "Interaction Optimizer Real-time Adherence Tracking",
    "Interaction Optimizer Schedulable",
    ]

roles = ["MKT-Agent", "MKT-SF-Agent", "MKT-CC-Agent"]

column_header = {}

#app.dlg.print_control_identifiers() #Check Identifiers

class IAAutoGUI:
    def __init__(self, master, server = 0):
        self.server = server
        self.master = master
        master.title("IA Auto")
        self.master.minsize(width=300,height=400)

        #self.browse_button = ""
        #self.browse_button.grid(row = 1)

        self.filename = fd.askopenfilename()
        self.filename_label = Label(master, text = self.filename)
        self.filename_label.grid(row = 1, sticky = W)

        self.wb = load_workbook(self.filename, read_only=True)
        self.sh = self.wb.worksheets[0] # Change Sheets
        self.ws = self.wb.active

        
        self.serv_label = Label(master, text="Select Server")
        self.serv_label.grid(columnspan = 2, sticky = W)
        
        self.cms_button = Button(master, text="CMS", command=lambda:self.get_server(1))
        self.cms_button.grid(row = 2)

        self.sf_button = Button(master, text="Salesforce", command=lambda:self.get_server(2))
        self.sf_button.grid(row = 2, column = 1, sticky = W)

        self.loc_label = Label(master, text="Select Location")
        self.loc_label.grid(row = 3, columnspan = 2, sticky = W, pady = 10)

        self.orl_button = Button(master, text="orl", command=lambda:self.get_location("orl"),
                                 state=DISABLED, width = 20, padx = 20)
        self.orl_button.grid(row = 4)

        self.spg_button = Button(master,text="spg", command=lambda:self.get_location("spg"),
                                 state=DISABLED, width = 20, padx = 20)
        self.spg_button.grid(row = 5)  
        self.lvn_button = Button(master, text="lvn", command=lambda:self.get_location("lvn"),
                                 state=DISABLED, width = 20, padx = 20)
        self.lvn_button.grid(row = 4, column = 2, sticky = W )
        
    

        self.dept_label = Label(master, text="Select Department")
        self.dept_label.grid(row = 6, sticky = W, pady = 10)

        self.department = StringVar()
        self.outbnd = Radiobutton(master,indicatoron = 0, width = 20, padx = 20,
                                  text='Outbound', variable=self.department, value = "outbnd")
        self.outbnd.grid(row = 7)

        self.ct = Radiobutton(master, indicatoron = 0, width = 20, padx = 20,
                              text='Call Transfer', variable=self.department, value = "ct")
        self.ct.grid(row = 7, column = 2)
        
        self.act = Radiobutton(master, indicatoron = 0, width = 20, padx = 20,
                               text='Activations', variable=self.department, value = "act")
        self.act.grid(row = 8)
        
        self.cc = Radiobutton(master, indicatoron = 0, width = 20, padx = 20,
                              text='Customer Care', variable=self.department, value = "cc")
        self.cc.grid(row = 8, column = 2)

        self.run_button =  Button(master, width = 10, height = 2, padx = 10,
                                  text = "Run", command = lambda:self.get_department())
        self.run_button.grid(row = 10, pady = 10)
        
        #self.reset_button =  Button(master, width = 10, height = 2, padx = 10, text = "Reset", command = lambda:self.restart_button())
        #self.reset_button.grid(row = 9, column =  2, pady = 10)

        col_count, row_count = root.grid_size()
        
    def get_server(self, button_id):
        if button_id == 1:
            self.server = 1
            print("cms")
        elif button_id == 2:
            self.server = 2
            print("salesforce")
            
        self.cms_button.configure(state=DISABLED)
        self.sf_button.configure(state=DISABLED)
        self.orl_button.configure(state=NORMAL)
        self.lvn_button.configure(state=NORMAL)
        self.spg_button.configure(state=NORMAL)
    
    def get_location(self, location):
        self.location = location
        
        print(location)
        self.orl_button.configure(state=DISABLED)
        self.lvn_button.configure(state=DISABLED)
        self.spg_button.configure(state=DISABLED)

        if self.server == 1:
            if location == "spg":
                self.outbnd.configure(state = DISABLED)
                self.act.configure(state = DISABLED)
                self.cc.configure(state = DISABLED)
            elif location == "lvn":
                self.act.configure(state = DISABLED)
                self.cc.configure(state = DISABLED)
            else:
                pass
        else:
            self.ct.configure(state = DISABLED)
            self.act.configure(state = DISABLED)
            self.cc.configure(state = DISABLED)
            self.act.configure(state = DISABLED)
            self.cc.configure(state = DISABLED)
            
    def get_department(self):
        self.department = self.department.get()
        print(self.department)
        
        self.outbnd.configure(state = DISABLED)
        self.ct.configure(state = DISABLED)
        self.act.configure(state = DISABLED)
        self.cc.configure(state = DISABLED)
        self.act.configure(state = DISABLED)
        self.cc.configure(state = DISABLED)
        
        IAserver()
        main()
 
    def restart_button(self):
        pass

app = Application(backend='uia')
def IAserver():
    print(my_gui.server)
    if my_gui.server == 1:
        p = pywinauto.findwindows.find_element(title="Interaction Administrator - [HiltonACD]")
        app.connect(handle=p.handle)
        dlg = app.window(title="Interaction Administrator - [HiltonACD]")
    else:
        p = pywinauto.findwindows.find_element(title="Interaction Administrator - [HiltonTCPA]")
        app.connect(handle=p.handle)
        dlg = app.window(title="Interaction Administrator - [HiltonTCPA]")
        
def main():
    """Main Function"""
    get_alphabet()
    column_headers()

def get_alphabet():
    """Loop Through Alphabet and Column Headers in Excel"""
    for c in ascii_lowercase:
        _x = my_gui.sh[c.upper() + "1"].value
        column_header[c.upper()] = my_gui.sh[c.upper() + "1"].value

def get_header(getLetter):
    """Match Header of File"""
    for k, v in column_header.items():
        if v == getLetter:
            getLetter = k
            return str(k)
        
def column_headers():
    """Assign ColumnHeader Names"""
    Add = "Add/Delete/Change/Transfer/Rehire"
    Add2 = "Add/Delete/Change"
    Add3 = "Add/Deactivate/Reactivate/Change"
    Add4 = "Add/Delete/Change/Transfer"
    windows = "Windows Username"
    windows2 = "Windows for Adds"
    Name = "AgentName"
    cic_id = "CIC_ID"
    cic_id2 = "CIC ID"

    add = get_header(Add) or get_header(Add2) or get_header(Add3) or get_header(Add4)
    agent_username = get_header(windows) or get_header(windows2) 
    agent_name = get_header(Name)
    agent_tsr = get_header(cic_id) or get_header(cic_id2)

    orgchart_data(add, agent_username, agent_name, agent_tsr)


def orgchart_data(add, windows, agent_name, agent_tsr):
    n = 2
    while n < my_gui.sh.max_row + 1:
        if my_gui.sh[add + str(n)].value == "Add":
            username = my_gui.sh[windows + str(n)].value
            agentName = my_gui.sh[agent_name + str(n)].value
            tsr = my_gui.sh[agent_tsr + str(n)].value

            print("Adding User - " + username, agentName, tsr)

            try:
                config(tsr, username)
                get_user_details(agentName)
                auto_acd()
                if my_gui.server == 1:
                    cms_roles(my_gui.department)
                    get_cms_workgroups(my_gui.location, my_gui.department)
                    if my_gui.department == "ct" or my_gui.department == "cc":
                        licensing()
                    app.dlg.Cancel.click_input() #Change When Done
                else:
                    sf_roles()
                    get_sf_workgroups()
                    app.dlg.Cancel.click_input() #Change When Done
            except:
                if app.dlg["A User with that name already exists"].exists() == True:
                    print("User Already Exists " + username, agentName, tsr)
                    app.dlg.OK.click_input()
                    app.dlg.Cancel.click_input()
        n += 1

def config(tsr, win_username, passwd="10102015", domain="hgvcnt\\"):
    """Assign Configuration"""
    app.dlg.type_keys('^n')
    app.dlg.Edit0.type_keys(str(tsr) + "{ENTER}")
    app.dlg.Edit3.type_keys(passwd) #Password
    app.dlg.Edit4.type_keys(passwd) #Confirm Password
    app.dlg.Edit7.type_keys(domain + win_username) #Domain User
    get_location()

def get_location():
    """Gets Location from combobox"""
    app.dlg["ComboBox4"].click_input()
    if my_gui.location == "orl":
        app.dlg["Orlando - Metro Center"].select()
    if my_gui.location == "spg":
        app.dlg["Spingfield"].select()
    if my_gui.location == "lvn":
        app.dlg["Las Vegas"].select()

def get_user_details(agentName):
    """Assign Personal Info"""
    name = agentName.split(" ")

    app.dlg.TabItem3.click_input()
    app.dlg.Edit0.type_keys(name[0]) #FirstName
    app.dlg.Edit2.type_keys(name[1]) #LastName
    app.dlg.Edit4.type_keys(name[0] + "{SPACE}" + name[1]) #Display Name

def auto_acd():
    """Assign Auto-ACD"""
    app.dlg.ACD.click_input()
    app.dlg.ListItem3.click_input()
    app.dlg.CheckBox0.click_input()

def cms_roles(deptmnt):
    """Assign Roles for cms"""
    app.dlg.Roles.click_input()
    app.dlg.Add.click_input()

    if deptmnt == "ct" or deptmnt == "outbnd":
        app.dlg["MKT-Agent"].select()
    elif deptmnt == "act" or deptmnt == "cc":
        app.dlg["MKT-CC-Agent"].select()

    app.dlg.OK.click_input()

def sf_roles():
    """Assign Roles in SalesForce"""
    app.dlg.Roles.click_input()
    app.dlg.Add.click_input()
    app.dlg.ListItem6.select()
    app.dlg.OK.click_input()

def get_cms_workgroups(loc, dept):
    """Assign CMS WorkGroups based on department"""
    if my_gui.location == loc:
        if my_gui.department == dept:
            agent_cms_workgroups(wrkqueues[loc + "_" + dept + "_" + "cms"])

def get_sf_workgroups(): #helper function here
    """Assign Salesforce WorkGroups based on Selection"""
    if my_gui.location == "orl":
        agent_sf_workgroups(orl_outbnd_sf)
    if my_gui.location == "spg":
        agent_sf_workgroups(spg_outbnd_sf)
    if my_gui.location == "lvn":
        agent_sf_workgroups(lvn_outbnd_sf)

def agent_cms_workgroups(wrkgrps):
    """Assign CMS WorkGroups"""
    app.dlg.Workgroups.click_input()
    app.dlg.OK.click_input()
    
    for x in wrkgrps:
        if my_gui.location == "orl":
            if my_gui.department == "ct":
                print(wrkgrps)
                LstBoxAdd(x)

            if my_gui.department == "act":
                LstBoxAdd(x)
            if my_gui.department == "cc":
                LstBoxAdd(x)

        if my_gui.location == "spg":
            LstBoxAdd(x)
            
        if my_gui.location == "lvn":
            if my_gui.department == "outbnd":
                LstBoxAdd(x)
            if my_gui.department == "ct":
                LstBoxAdd(x)

def LstBoxAdd(x):
    app.dlg['ListBox'].click_input()
    app.dlg['Listbox'].type_keys(x)
    app.dlg.Add.click_input()

def agent_sf_workgroups(wrkgrps):
    """Assign SF WorkGroups"""
    num = 0
    app.dlg.Workgroups.click_input()
    app.dlg.OK.click_input()

    for x in wrkgrps:
        if my_gui.location == "orl":
            LstBoxAdd(x)
        if my_gui.location == "spg":
            LstBoxAdd(x)
        if my_gui.location == "lvn":
            LstBoxAdd(x)

def listbox_pos(scrollpos):
    """Gets Scrollbox Position"""
    app.dlg['ListBox'].click_input()
    app.dlg['ListBox'].wheel_mouse_input(wheel_dist=scrollpos)

def licensing():
    """Assign Licensing"""
    app.dlg.Licensing.click_input()
    app.dlg.OK.click_input()
    app.dlg['Enable Licenses'].click_input()
    for ls in licenses:
        app.dlg[ls].type_keys("{SPACE}")


if __name__ == '__main__':
    root = Tk()
    my_gui = IAAutoGUI(root)
    root.mainloop()
    
sys.exit()
